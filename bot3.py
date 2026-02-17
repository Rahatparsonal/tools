# Decompiled with PyLingual (https://pylingual.io)
# Internal filename: 'forget.py'
# Bytecode version: 3.11a7e (3495)
# Source timestamp: 1970-01-01 00:00:00 UTC (0)

global total_success
global settings
global total_checked
global SELECTED_SERVER
global CURRENT_LOCALE
global total_error
global SELECTED_DEVICE
global user_nm
global total_failed
global expr
global SELECTED_BROWSER
# ***<module>: Failure: Compilation Error
import os
import random
import re
import sys
import time
import platform
import webbrowser
import json
import certifi
import threading
import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
import ssl
import socket
import base64
from datetime import datetime, timezone, timedelta
import wmi
import hashlib
import subprocess
import winreg
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
import itertools
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')
    os.system('')
if getattr(sys, 'frozen', False):
    os.chdir(os.path.dirname(sys.executable))
WHITE = '[1;97m'
GREEN = '[1;92m'
RED = '[1;91m'
DARK_GREEN = '[1;32m'
LIGHT_GRAY = '[1;37m'
CYAN = '[1;96m'
YELLOW = '[1;93m'
BLUE = '[1;94m'
MAGENTA = '[1;95m'
ORANGE = '[38;5;208m'
GOLD = '[38;5;220m'
VIOLET = '[38;5;141m'
TOXIC = '[38;2;170;200;0m'
PURPLE = '[38;2;150;80;200m'
RESET = '[0m'
opt_labels = [f'{GREEN}[{RED}{str(i).zfill(2)}{GREEN}]' for i in range(1, 8)]
l0 = f'{GREEN}[{RED}00{GREEN}]'
EKL = f'{CYAN}:{WHITE}'
LINE = f'{CYAN}•━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━•'
SERVER_URL = 'https://farhad80715.pythonanywhere.com'
def get_windows_device_id():
    """Windows Device ID Generate করে"""
    # ***<module>.get_windows_device_id: Failure detected at line number 10 and instruction offset 6: Different bytecode
    unique_factors = []
    try:
        key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, 'SOFTWARE\\Microsoft\\Cryptography', 0, winreg.KEY_READ | winreg.KEY_WOW64_64KEY)
        guid, _ = winreg.QueryValueEx(key, 'MachineGuid')
        winreg.CloseKey(key)
        if guid:
            unique_factors.append(str(guid))
    except:
        pass
    try:
        c = wmi.WMI()
        uuid = c.Win32_ComputerSystemProduct()[0].UUID
        if uuid:
            unique_factors.append(str(uuid))
    except:
        pass
    try:
        cmd = 'wmic csproduct get uuid'
        uuid_raw = subprocess.check_output(cmd, shell=True, stderr=subprocess.DEVNULL).decode('ignore', errors='ignore').strip()
        lines = [line.strip() for line in uuid_raw.split('\n') if line.strip()]
        if len(lines) > 1 and lines[1]:
                unique_factors.append(lines[1])
    except:
        pass
    if not unique_factors:
        try:
            hafky = os.getlogin().upper()
        except:
            hafky = 'USER'
        platfm = platform.version()
        unique_factors.append(hafky + platfm)
    raw_id = ''.join(unique_factors).replace(' ', '').replace('-', '').upper()
    hashed_id = hashlib.sha256(raw_id.encode()).hexdigest().upper()
    return hashed_id[:32]
def check_approval():
    """Server থেকে Approval Check করে"""
    global user_nm
    global expr
    # ***<module>.check_approval: Failure detected at line number 48 and instruction offset 8: Different bytecode
    try:
        device_id = get_windows_device_id()
        response = requests.get(f'{SERVER_URL}/apv', timeout=10)
        if response.status_code == 200:
            devices = response.json()
            for device in devices:
                if device['Device_ID'] == device_id:
                    user_nm = device['User_Name']
                    expr = device['End_date']
                    nw_tm = datetime.now(timezone.utc)
                    expirs = datetime.strptime(expr, '%Y-%m-%d %H:%M').replace(tzinfo=timezone.utc)
                    if nw_tm >= expirs:
                        clear_logo()
                        print(f'\n{LINE}')
                        print(f'{RED}  YOUR ACCESS HAS EXPIRED!{RESET}')
                        print(f'{LINE}')
                        print(f'{WHITE}  Device ID  : {GREEN}{device_id}{RESET}')
                        print(f'{WHITE}  User Name  : {GREEN}{user_nm}{RESET}')
                        print(f'{WHITE}  Expired on : {RED}{expr} (UTC){RESET}')
                        print(f'{LINE}\n')
                        input(f'{WHITE}  Press Enter to contact owner...{RESET}')
                        webbrowser.open('https://t.me/farhad80715')
                        sys.exit(0)
                    return True
            clear_logo()
            print(f'\n{LINE}')
            print(f'{RED}  DEVICE NOT REGISTERED!{RESET}')
            print(f'{LINE}')
            print(f'{WHITE}  Device ID : {YELLOW}{device_id}{RESET}')
            print(f'{WHITE}  Your Device ID is not registered.{RESET}')
            print(f'{WHITE}  Please contact owner to get access.{RESET}')
            print(f'{LINE}\n')
            input(f'{WHITE}  Press Enter to contact owner...{RESET}')
            webbrowser.open('https://t.me/farhad80715')
            sys.exit(0)
        else:
            clear_logo()
            print(f'\n{LINE}')
            print(f'{RED}  SERVER CONNECTION FAILED!{RESET}')
            print(f'{LINE}')
            print(f'{WHITE}  Please check your internet connection.{RESET}')
            print(f'{WHITE}  Server URL: {CYAN}{SERVER_URL}{RESET}')
            print(f'{LINE}\n')
            input(f'{WHITE}  Press Enter to exit...{RESET}')
            sys.exit(0)
    except requests.exceptions.ConnectionError:
        clear_logo()
        print(f'\n{LINE}')
        print(f'{RED}  SERVER OFFLINE!{RESET}')
        print(f'{LINE}')
        print(f'{WHITE}  Cannot connect to approval server.{RESET}')
        print(f'{WHITE}  Server URL: {CYAN}{SERVER_URL}{RESET}')
        print(f'{LINE}\n')
        input(f'{WHITE}  Press Enter to exit...{RESET}')
        sys.exit(0)
    except Exception as e:
        clear_logo()
        print(f'\n{LINE}')
        print(f'{RED}  APPROVAL ERROR!{RESET}')
        print(f'{LINE}')
        print(f'{WHITE}  Error: {RED}{e}{RESET}')
        print(f'{LINE}\n')
        input(f'{WHITE}  Press Enter to exit...{RESET}')
        sys.exit(0)
SERVER_MAP = {1: 'm.facebook.com', 2: 'mbasic.facebook.com', 3: 'touch.facebook.com', 4: 'free.facebook.com', 5: 'm.alpha.facebook.com', 6: 'm.beta.facebook.com', 7: 'x.facebook.com', 8: 'limited.facebook.com'}
DEVICE_MAP = {1: 'Android', 2: 'iPhone', 3: 'KaiOS', 4: 'Windows Phone', 5: 'BlackBerry'}
print_lock = threading.Lock()
counter_lock = threading.Lock()
total_checked = 0
total_success = 0
total_failed = 0
total_error = 0
PROXIES = None
CURRENT_LOCALE = 'en_US'
SELECTED_SERVER = 'limited.facebook.com'
SELECTED_DEVICE = 'Android'
SELECTED_BROWSER = 'Chrome'
user_nm = 'Unknown'
expr = 'Unknown'
settings = {}
def make_request(url):
    # ***<module>.make_request: Failure: Different control flow
    if url.startswith('https://'):
        url = url[8:]
    else:
        if url.startswith('http://'):
            url = url[7:]
    if '/' in url:
        host, path = url.split('/', 1)
        path = '/' + path
    else:
        host, path = (url, '')
    context = ssl.create_default_context(cafile=certifi.where())
    context.check_hostname = True
    context.verify_mode = ssl.CERT_REQUIRED
    try:
        with socket.create_connection((host, 443)) as sock:
            with context.wrap_socket(sock, server_hostname=host) as ssock:
                request = f'GET {path} HTTP/1.1\r\n'
                request += f'Host: {host}\r\n'
                request += 'User-Agent: Python-Socket\r\n'
                request += 'Accept: */*\r\n'
                request += 'Connection: close\r\n\r\n'
                ssock.sendall(request.encode())
                response = b''
                while True:
                    data = ssock.recv(4096)
                    if not data:
                        pass
                    else:
                        response += data
    except Exception:
        return None
    try:
        header_data, body_data = response.split(b'\r\n\r\n', 1)
        headers = header_data.decode('ignore', errors='ignore').split('\r\n')
        status_line = headers[0]
        status_code = int(status_line.split()[1])
        headers_dict = {}
        cookies = {}
        for header in headers[1:]:
            key, value = header.split(':', 1)
            headers_dict[key.strip()] = value.strip()
            if key.lower() == 'set-cookie':
                cookie_parts = value.split(';')[0].split('=')
                cookies[cookie_parts[0].strip()] = cookie_parts[1].strip()
        response_text = body_data.decode('ignore', errors='ignore')
        try:
            response_json = json.loads(response_text)
        except json.JSONDecodeError:
            response_json = None
        return {'status_code': status_code, 'text': response_text, 'json': response_json, 'headers': headers_dict, 'cookies': cookies}
    except Exception:
        return None
SECRET_KEY = b'LHANKLRTOLUMCDCK'
SECRET_KEY2 = b'GTRMAREAMLXUDWDJ'
def dec_rq(sxrreqq):
    # ***<module>.dec_rq: Failure detected at line number 164 and instruction offset 2: Different bytecode
    dec_base4 = base64.urlsafe_b64decode(sxrreqq.encode('utf-8'))
    cipher = AES.new(SECRET_KEY, AES.MODE_ECB)
    dec_cryoto = unpad(cipher.decrypt(dec_base4), AES.block_size).decode('utf-8')
    jsn_dta = json.loads(dec_cryoto)
    return jsn_dta
def dec_rq2(keyid):
    # ***<module>.dec_rq2: Failure detected at line number 170 and instruction offset 2: Different bytecode
    c_base4 = base64.urlsafe_b64decode(keyid.encode('utf-8'))
    ciphr = AES.new(SECRET_KEY2, AES.MODE_ECB)
    c_cryoto = unpad(ciphr.decrypt(c_base4), AES.block_size).decode('utf-8')
    return c_cryoto
def get_safe_cmd(cmd):
    # ***<module>.get_safe_cmd: Failure detected at line number 176 and instruction offset 2: Different bytecode
    try:
        return subprocess.check_output(cmd, shell=True, stderr=subprocess.DEVNULL).decode('ignore', errors='ignore').strip()
    except:
        return None
def get_windows_device_id_old():
    # ***<module>.get_windows_device_id_old: Failure detected at line number 182 and instruction offset 6: Different bytecode
    unique_factors = []
    try:
        key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, 'SOFTWARE\\Microsoft\\Cryptography', 0, winreg.KEY_READ | winreg.KEY_WOW64_64KEY)
        guid, _ = winreg.QueryValueEx(key, 'MachineGuid')
        winreg.CloseKey(key)
        if guid:
            unique_factors.append(str(guid))
    except Exception:
        pass
    uuid_found = False
    try:
        cmd = 'wmic csproduct get uuid'
        uuid_raw = subprocess.check_output(cmd, shell=True, stderr=subprocess.DEVNULL).decode('ignore', errors='ignore').strip()
        lines = [line.strip() for line in uuid_raw.split('\n') if line.strip()]
        if len(lines) > 1 and lines[1]:
                unique_factors.append(lines[1])
                uuid_found = True
    except Exception:
        pass
    if not uuid_found:
        try:
            c = wmi.WMI()
            uuid = c.Win32_ComputerSystemProduct()[0].UUID
            if uuid:
                unique_factors.append(str(uuid))
                uuid_found = True
        except Exception:
            pass
    if not uuid_found:
        try:
            ps_uuid = get_safe_cmd('powershell -Command \"Get-WmiObject -Class Win32_ComputerSystemProduct | Select-Object -ExpandProperty UUID\"')
            if ps_uuid:
                unique_factors.append(ps_uuid)
                uuid_found = True
        except Exception:
            pass
    hw_info_found = False
    try:
        c = wmi.WMI()
        baseboard_id = c.Win32_BaseBoard()[0].SerialNumber
        processor_id = c.Win32_Processor()[0].ProcessorId
        unique_factors.append(f'{baseboard_id}-{processor_id}')
        hw_info_found = True
    except Exception:
        pass
    if not hw_info_found:
        try:
            bb_id = get_safe_cmd('powershell -Command \"Get-WmiObject -Class Win32_BaseBoard | Select-Object -ExpandProperty SerialNumber\"')
            proc_id = get_safe_cmd('powershell -Command \"Get-WmiObject -Class Win32_Processor | Select-Object -ExpandProperty ProcessorId\"')
            if not bb_id and (not proc_id):
                pass
            else:
                if not bb_id:
                    bb_id = 'None'
                if not proc_id:
                    proc_id = 'None'
                unique_factors.append(f'{bb_id}-{proc_id}')
                hw_info_found = True
        except Exception:
            pass
    if not unique_factors:
        try:
            hafky = os.getlogin().upper()
        except:
            hafky = 'GENERIC_USER'
        platfm = platform.version()
        unique_factors.append(hafky + platfm)
    raw_id = ''.join(unique_factors).replace(' ', '').replace('-', '').upper()
    hashed_id = hashlib.sha256(raw_id.encode()).hexdigest().upper()
    return hashed_id[:32]
def apvv():
    sxr_main()
def parse_proxy(proxy_str):
    if '://' not in proxy_str:
        parts = proxy_str.split(':')
        if len(parts) == 4:
            ip, port, user, pwd = parts
            proxy_url = f'http://{user}:{pwd}@{ip}:{port}'
        else:
            if len(parts) == 2:
                ip, port = parts
                proxy_url = f'http://{ip}:{port}'
            else:
                return None
    else:
        proxy_url = proxy_str
    return {'http': proxy_url, 'https': proxy_url}
def test_proxy(proxies, server_domain):
    # ***<module>.test_proxy: Failure detected at line number 269 and instruction offset 2: Different bytecode
    try:
        r = requests.get(f'https://{server_domain}', proxies=proxies, timeout=10)
        return r.status_code == 200
    except:
        return False
sv_SE = {'AD': 'ca_ES', 'AE': 'ar_AR', 'AF': 'fa_IR', 'AG': 'en_US', 'AI': 'en_US', 'AL': 'sq_AL', 'AM': 'hy_AM', 'sq_AL': 'es_LA', 'hy_AM': 'ar_AR', 'AO': 'es_LA', 'pt_PT': 'en_US', 'AQ': 'ar_AR', 'AR': 'es_LA', 'es_LA': 'ar_AR', 'AI': 'en_US', 'AL': 'es_LA', 'sq_AL': 'en_US', 'AM': 'ar_AR', 'hy_AM': 'es_LA', 'AO': 'ar_AR', 'pt_PT': 'en_US', 'AQ': 'es_LA', 'AR': 'es_LA', 'es_LA': 'ar_AR', 'AS': 'en_US',
def get_locale_code(country_code):
    return COUNTRY_TO_LOCALE.get(country_code.upper(), 'en_US')
def get_ip_info(proxies=None):
    # irreducible cflow, using cdg fallback
    # ***<module>.get_ip_info: Failure: Compilation Error
    r = requests.get('http://ip-api.com/json/', proxies=proxies, timeout=10)
    if r.status_code == 200:
        data = r.json()
        return {'country': data.get('country', 'Unknown'), 'countryCode': data.get('countryCode', 'US'), 'timezone': data.get('timezone', 'Unknown')}
                pass
                return {'country': 'Unknown', 'countryCode': 'US', 'timezone': 'Unknown'}
def load_settings():
    # irreducible cflow, using cdg fallback
    # ***<module>.load_settings: Failure: Compilation Error
    with open('Setting.json', 'r') as f:
        return json.load(f)
                return {}
def get_status_line():
    return f'\r{GREEN}[{WHITE}Farhad Tools{GREEN}] {WHITE}CHECKED:-{total_checked}{CYAN}|{GREEN}SUCCESS:-{total_success}{CYAN}|{YELLOW}FAILED:-{total_failed}{CYAN}|{RED}ERROR:-{total_error}'
def safe_print(text):
    with print_lock:
        sys.stdout.write('\r                                                                                \r')
        try:
            sys.stdout.write(str(text) + '\n')
        except UnicodeEncodeError:
            sys.stdout.write(str(text).encode('utf-8', 'ignore').decode('utf-8') + '\n')
        sys.stdout.write(get_status_line())
        sys.stdout.flush()
def update_counter(status, number=None, message=None, color=None, html_content=None):
    global total_failed
    global total_success
    global total_checked
    global total_error
    with counter_lock:
        if status == 'success':
            total_success += 1
        else:
            if status == 'failed':
                total_failed += 1
            else:
                if status == 'error':
                    total_error += 1
                    if html_content:
                        save_error_html(message if message else 'Unknown Error', html_content)
        total_checked += 1
    if message and number:
        if not color:
            color = WHITE
        safe_print(f'{color} {message} {number}')
    else:
        if message:
            if not color:
                color = WHITE
            safe_print(f'{color} {message}')
        else:
            with print_lock:
                sys.stdout.write(get_status_line())
                sys.stdout.flush()
SAVE_ERROR_LOGS = 'off'
def reset_counters():
    global total_failed
    global total_success
    global total_checked
    global total_error
    total_checked = 0
    total_success = 0
    total_failed = 0
    total_error = 0
def save_error_html(message, html_content):
    # irreducible cflow, using cdg fallback
    # ***<module>.save_error_html: Failure: Compilation Error
    if SAVE_ERROR_LOGS.lower()!= 'on':
        return None
    if not os.path.exists('Error_Logs'):
        os.makedirs('Error_Logs')
    safe_msg = re.sub('[\\\\/*?:\"<>|]', '', message)
    safe_msg = safe_msg.replace(' ', '_')
    safe_msg = safe_msg[:50]
    base_filename = f'Error_Logs/{safe_msg}.html'
    filename = base_filename
    counter = 1
    while os.path.exists(filename):
        filename = f'Error_Logs/{safe_msg}_{counter}.html'
        counter += 1
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f'<!-- Error: {message} -->\n')
        f.write(html_content)
                except Exception as e:
                        safe_print(f'{RED} Failed to save error log: {e}')
def clear_logo():
    # ***<module>.clear_logo: Failure: Different bytecode
    if platform.system() == 'Windows':
        os.system('cls')
    else:
        os.system('clear')
    logo = '\n'.join(f'{GREEN}/$$$$$$$$ /$$$$$$  /$$$$$$$  /$$   /$$  /$$$$$$  /$$$$$$$        /$$$$$$$$ /$$$$$$   /$$$$$$  /$$        /$$$$$$ \n| $$_____//$$__  $$| $$__  $$| $$  | $$ /$$__  $$| $$__  $$      |__  $$__//$$__  $$ /$$__  $$| $$       /$$__  $$\n| $$     | $$  \\ $$| $$  \\ $$| $$  | $$| $$  \\ $$| $$  \\ $$         | $$  | $$  \\ $$| $$  \\ $$| $$      | $$  \\__/\n| $$$$$  | $$$$$$$$| $$$$$$$/| $$$$$$$$| $$$$$$$$| $$  | $$         | $$  | $$  | $$| $$  | $$| $$      |  $$$$$$ \n| $$__/  | $$__  $$| $$__  $$| $$__  $$| $$__  $$| $$  | $$         | $$  | $$  | $$| $$  | $$| $$       \\____  $$\n| $$     | $$  | $$| $$  \\ $$| $$  | $$| $$  | $$| $$  | $$         | $$  | $$  | $$| $$  | $$| $$       /$$  \\ $$\n| $$     | $$  | $$| $$  | $$| $$  | $$| $$  | $$| $$$$$$$/         | $$  |  $$$$$$/|  $$$$$$/| $$$$$$$$|  $$$$$$/\n|__/     |__/  |__/|__/  |__/|__/  |__/|__/  |__/|_______/          |__/   \\______/  \\______/ |________/ \\______/ \n{LINE}\n {GREEN}[{RED}●{GREEN}] TOOL OWNER   {CYAN}:{GREEN} FARHAD AHMED\n {GREEN}] TOOL         {CYAN}:{GREEN} FORGET FB\n {GREEN}] TOOL STATUS  {CYAN}:{GREEN} PAID\n {GREEN}] USER NAME    {CYAN}:{GREEN} {user_nm}\n {GREEN}:{GREEN} {expr}\n{LINE}')
    print(logo)
def select_server():
    global SELECTED_SERVER
    # ***<module>.select_server: Failure detected at line number 351 and instruction offset 210: Different bytecode
    clear_logo()
    print(f'\n {GREEN}SERVER SELECT:\n{LINE}')
    for i in range(1, 9):
        print(f' {GREEN}[{RED}{str(i).zfill(2)}{GREEN}] {SERVER_MAP[i]}')
    print(LINE)
    while True:
        try:
            server_choice = input(f'\n{GREEN} [{RED}●{GREEN}] Select Server (1-8) {EKL} ').strip()
            if server_choice in [str(i) for i in range(1, 9)]:
                SELECTED_SERVER = SERVER_MAP[int(server_choice)]
                print(f'\n{GREEN} [{RED}●{GREEN}] Selected Server {EKL} {SELECTED_SERVER}')
                time.sleep(1)
                return
            else:
                print(f'{RED} Invalid selection! Please enter 1-8')
        except:
            print(f'{RED} Invalid input!')
def select_device():
    global SELECTED_DEVICE
    # ***<module>.select_device: Failure detected at line number 370 and instruction offset 254: Different bytecode
    clear_logo()
    print(f'\n {GREEN}DEVICE SELECT:\n{LINE}')
    print(f' {GREEN}[{RED}01{GREEN}] Android (Default)    {GREEN}[{RED}02{GREEN}] iPhone')
    print(f' {GREEN}[{RED}03{GREEN}] KaiOS                {GREEN}[{RED}04{GREEN}] Windows Phone')
    print(f' {GREEN}[{RED}05{GREEN}] BlackBerry\n{LINE}')
    while True:
        try:
            device_choice = input(f'\n{GREEN} [{RED}●{GREEN}] Select Device {EKL} ').strip().zfill(2)
            if device_choice in ['01', '02', '03', '04', '05']:
                SELECTED_DEVICE = DEVICE_MAP[int(device_choice)]
                print(f'\n{GREEN} [{RED}●{GREEN}] Selected Device {EKL} {SELECTED_DEVICE}')
                time.sleep(1)
                return
            else:
                print(f'{RED} Invalid selection! Please enter 1-5')
        except:
            print(f'{RED} Invalid input!')
def select_browser():
    global SELECTED_BROWSER
    # ***<module>.select_browser: Failure detected at line number 403 and instruction offset 678: Different bytecode
    clear_logo()
    print(f'\n {GREEN}BROWSER SELECT:\n{LINE}')
    print(f' {GREEN}[{RED}01{GREEN}] Chrome')
    print(f' {GREEN}[{RED}02{GREEN}] Firefox')
    print(f' {GREEN}[{RED}03{GREEN}] Opera')
    print(f' {GREEN}[{RED}04{GREEN}] Edge')
    print(f' {GREEN}[{RED}05{GREEN}] Brave')
    print(f' {GREEN}[{RED}06{GREEN}] Samsung')
    print(f' {GREEN}[{RED}07{GREEN}] UC Browser')
    print(f' {GREEN}[{RED}08{GREEN}] DuckDuckGo')
    print(f' {GREEN}[{RED}09{GREEN}] Vivaldi')
    print(f' {GREEN}[{RED}10{GREEN}] Yandex')
    print(f' {GREEN}[{RED}11{GREEN}] Kiwi')
    print(f' {GREEN}[{RED}12{GREEN}] Dolphin')
    print(f' {GREEN}[{RED}13{GREEN}] Mi Browser')
    print(f' {GREEN}[{RED}14{GREEN}] Maxthon')
    print(f' {GREEN}[{RED}15{GREEN}] Puffin')
    print(f' {GREEN}[{RED}00{GREEN}] Random (Mix)\n{LINE}')
    browser_map = {'01': 'Chrome', '02': 'Firefox', '03': 'Opera', '04': 'Edge', '05': 'Brave', '06': 'Samsung', '07': 'UC', '08': 'DuckDuckGo', '09': 'Vivaldi', '10': 'Yandex', '11': 'Kiwi', '12': 'Dolphin', '13': 'Mi Browser', '14': 'Maxthon', '15': 'Puffin', '00': 'Random'}
    while True:
        try:
            browser_choice = input(f'\n{GREEN} [{RED}●{GREEN}] Select Browser {EKL} ').strip().zfill(2)
            if browser_choice in browser_map:
                SELECTED_BROWSER = browser_map[browser_choice]
                print(f'\n{GREEN} [{RED}●{GREEN}] Selected Browser {EKL} {SELECTED_BROWSER}')
                time.sleep(1)
                return
            else:
                print(f'{RED} Invalid selection! Please enter 00-15')
        except:
            print(f'{RED} Invalid input!')
def get_proxy_list(settings_key, prompt_label):
    # irreducible cflow, using cdg fallback
    # ***<module>.get_proxy_list: Failure: Compilation Error
    settings = load_settings()
    proxy_set = settings.get(settings_key, {})
    ask_proxy = proxy_set.get('ask_for_proxy', True)
    def_proxy = proxy_set.get('default_proxy', '')
    server_domain = SELECTED_SERVER
    ask_proxy_final = ask_proxy
    PROXY_LIST = []
    if def_proxy:
        if isinstance(def_proxy, list):
            print(f'{WHITE} Testing {len(def_proxy)} Default {prompt_label}...')
            for p in def_proxy:
                parsed = parse_proxy(p)
                if parsed and test_proxy(parsed, server_domain):
                    nfo = get_ip_info(parsed)
                    loc = get_locale_code(nfo['countryCode'])
                    PROXY_LIST.append({'proxy': parsed, 'locale': loc, 'country': nfo['country']})
                    print(f"{GREEN} [{RED}●{GREEN}] {prompt_label} Location {EKL} {nfo['country']}")
                    print(f'{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}')
                else:
                    print(f'{RED} Default {prompt_label} Connection Failed: {p}')
        else:
            parsed_proxies = parse_proxy(def_proxy)
            if parsed_proxies:
                print(f'{WHITE} Testing Default {prompt_label}...')
                if test_proxy(parsed_proxies, server_domain):
                    nfo = get_ip_info(parsed_proxies)
                    loc = get_locale_code(nfo['countryCode'])
                    PROXY_LIST.append({'proxy': parsed_proxies, 'locale': loc, 'country': nfo['country']})
                    print(f"{GREEN} [{RED}●{GREEN}] {prompt_label} Location {EKL} {nfo['country']}")
                    print(f'{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}')
                else:
                    print(f'{RED} Default {prompt_label} Connection Failed!')
            else:
                print(f'{RED} Invalid Default {prompt_label} Format!')
        if def_proxy and (not PROXY_LIST):
                print(f'{RED} All Default {prompt_label} Failed!')
                ask_proxy_final = True
    if PROXY_LIST and ask_proxy:
            ask_proxy_final = False
    if ask_proxy_final:
        proxy_input = input(f'{GREEN} [{RED}●{GREEN}] Enter {prompt_label} (or \'y\' for multiple) [Press Enter to Skip] {EKL} ').strip()
        if not proxy_input:
            print(f'{YELLOW} No proxy entered, skipping proxy configuration...')
                break
            if proxy_input.lower() == 'y':
                cnt_in = input(f'{GREEN} [{RED}●{GREEN}] How many {prompt_label}? {EKL} ')
                if cnt_in.strip():
                    cnt = int(cnt_in)
                    for i in range(cnt):
                            p_in = input(f'{WHITE} [{RED}●{WHITE}] Enter {prompt_label} [{i + 1}/{cnt}] {EKL} ').strip()
                            if p_in:
                                print(f'{WHITE} Testing {prompt_label}...')
                                parsed = parse_proxy(p_in)
                                if parsed and test_proxy(parsed, server_domain):
                                    nfo = get_ip_info(parsed)
                                    loc = get_locale_code(nfo['countryCode'])
                                    print(f"{GREEN} [{RED}●{GREEN}] {prompt_label} Location {EKL} {nfo['country']}")
                                    print(f'{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}')
                                    PROXY_LIST.append({'proxy': parsed, 'locale': loc, 'country': nfo['country']})
                                    print(f'{RED} Connection Failed or Invalid Format!')
                    print(f'{RED} Invalid Number!')
                        break
                if proxy_input:
                    parsed_proxies = parse_proxy(proxy_input)
                    if parsed_proxies:
                        print(f'{WHITE} Testing {prompt_label}...')
                        if test_proxy(parsed_proxies, server_domain):
                            nfo = get_ip_info(parsed_proxies)
                            loc = get_locale_code(nfo['countryCode'])
                            print(f"{GREEN} [{RED}●{GREEN}] {prompt_label} Location {EKL} {nfo['country']}")
                            print(f'{GREEN} [{RED}●{GREEN}] Locale      {EKL} {loc}')
                            PROXY_LIST.append({'proxy': parsed_proxies, 'locale': loc, 'country': nfo['country']})
                                break
                            print(f'{RED} {prompt_label} Connection Failed!')
                        print(f'{RED} Invalid {prompt_label} Format!')
                                if PROXY_LIST or not ask_proxy_final:
                                        break
                print(f'{RED} Invalid Input')
                            return PROXY_LIST
def sxr_main():
    # ***<module>.sxr_main: Failure detected at line number 424 and instruction offset 158: Different bytecode
    clear_logo()
    print(f'\n {WHITE}Welcome to Farhad Forget FB Tool!\n{LINE}')
    print(f' {opt_labels[0]} FB FORGET')
    print(f' {opt_labels[1]} JOIN TELEGRAM\n{LINE}')
    chic_opsn = input(f'{GREEN} [{RED}●{GREEN}] CHOOSE OPTION {EKL} ').strip().zfill(2)
    if chic_opsn in ['01', '1', 'A', 'a']:
        select_server()
        select_device()
        select_browser()
        file_inp()
    else:
        if chic_opsn in ['02', '2', 'B', 'b']:
            webbrowser.open('https://t.me/farhad80715')
            sxr_main()
        else:
            print(f'\n{RED} You have selected the wrong option..')
            time.sleep(3)
            sxr_main()
def extract_numbers_from_excel(filename):
    # irreducible cflow, using cdg fallback
    # ***<module>.extract_numbers_from_excel: Failure: Compilation Error
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb.active
    target_col_idx = None
    max_matches = 0
    for col_idx in range(1, sheet.max_column + 1):
            match_count = 0
            for row_idx in range(2, min(22, sheet.max_row + 1)):
                    cell_val = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_val:
                        s_val = str(cell_val).strip()
                        s_cleaned = re.sub('[\\s\\-\\(\\)\\+]', '', s_val)
                        if s_cleaned.isdigit():
                            if 7 <= len(s_cleaned) <= 15:
                                    match_count += 1
                    if match_count > max_matches:
                        max_matches = match_count
                        target_col_idx = col_idx
            if target_col_idx is None:
                return (None, 'No phone number column found.')
                numbers = []
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=target_col_idx, max_col=target_col_idx, values_only=True):
                        val = row[0]
                        if val:
                            s_val = str(val).strip()
                            s_cleaned = re.sub('[\\s\\-\\(\\)\\+]', '', s_val)
                            if s_cleaned.isdigit():
                                if 7 <= len(s_cleaned) <= 15:
                                        numbers.append(s_cleaned)
                        return (numbers, None)
                            except Exception as e:
                                    return (None, str(e))
def file_inp():
    # irreducible cflow, using cdg fallback
    # ***<module>.file_inp: Failure: Compilation Error
    clear_logo()
    settings = load_settings()
    file_settings = settings.get('file_input_settings', {})
    always_use_txt = file_settings.get('always_use_txt', False)
    use_multiple_excel = file_settings.get('use_multiple_excel_files', False)
    if always_use_txt:
        if os.path.exists('Number_List.txt'):
            with open('Number_List.txt', 'r', encoding='utf-8', errors='ignore') as f:
                numbers = [line.strip() for line in f if line.strip()]
            if numbers:
                print(f'{GREEN} [{RED}●{GREEN}] Selected File {EKL} Number_List.txt')
                input(f'{WHITE} Press Enter to Start Forgetting {len(numbers)} Numbers...')
                autom_main()
            else:
                print(f'{WHITE} \'Number_List.txt\' file is empty.')
                input(f'{WHITE} Press Enter to return to main menu...')
                sxr_main()
        else:
            print(f'{WHITE} \'Number_List.txt\' file was not found.')
            input(f'{WHITE} Press Enter to return to main menu...')
            sxr_main()
    if use_multiple_excel:
        xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and (not f.startswith('~$'))]
        if xlsx_files:
            print(f'{GREEN} [{RED}●{GREEN}] Found {len(xlsx_files)} Excel Files.')
            all_numbers = []
            for f in xlsx_files:
                print(f'{WHITE} Extracting from {EKL} {f}...')
                nums, err = extract_numbers_from_excel(f)
                if nums:
                    all_numbers.extend(nums)
                    print(f'{GREEN}  -> Found {len(nums)} numbers.')
                else:
                    print(f'{RED}  -> Failed: {err}')
            if all_numbers:
                all_numbers = list(set(all_numbers))
                with open('Number_List.txt', 'w', encoding='utf-8', errors='ignore') as f:
                    for num in all_numbers:
                        f.write(num + '\n')
                print(f'\n{GREEN} [{RED}●{GREEN}] Total Unique Numbers Extracted {EKL} {len(all_numbers)}')
                print(f'{GREEN} [{RED}●{GREEN}] Saved to \'Number_List.txt\'\n')
                input(f'{WHITE} Press Enter to Start Forgetting {len(all_numbers)} Numbers...')
                autom_main()
            else:
                print(f'{RED} No valid numbers found in any Excel files.')
                input(f'{WHITE} Press Enter to return to main menu...')
                sxr_main()
        else:
            if os.path.exists('Number_List.txt'):
                with open('Number_List.txt', 'r', encoding='utf-8', errors='ignore') as f:
                    numbers = [line.strip() for line in f if line.strip()]
                if numbers:
                    print(f'{WHITE} No Excel files found, using Number_List.txt')
                    print(f'{GREEN} [{RED}●{GREEN}] Selected File {EKL} Number_List.txt')
                    input(f'{WHITE} Press Enter to Start Forgetting {len(numbers)} Numbers...')
                    autom_main()
                else:
                    print(f'{WHITE} No Excel files found and \'Number_List.txt\' is empty.')
                    input(f'{WHITE} Press Enter to return to main menu...')
                    sxr_main()
            else:
                print(f'{WHITE} No Excel files found and \'Number_List.txt\' not found.')
                input(f'{WHITE} Press Enter to return to main menu...')
                sxr_main()
    files = [f for f in os.listdir('.') if f.endswith('.xlsx') and (not f.startswith('~$'))]
    if not files:
        if os.path.exists('Number_List.txt'):
            with open('Number_List.txt', 'r', encoding='utf-8', errors='ignore') as f:
                numbers = [line.strip() for line in f if line.strip()]
            if numbers:
                print(f'{GREEN} [{RED}●{GREEN}] Selected File {EKL} Number_List.txt')
                input(f'{WHITE} Press Enter to Start Forgetting {len(numbers)} Numbers...')
                autom_main()
            else:
                print(f'{WHITE} No Excel files found or \'Number_List.txt\' file is empty.')
                input(f'{WHITE} Press Enter to return to main menu...')
                sxr_main()
        else:
            print(f'{WHITE} No Excel files found and \'Number_List.txt\' were not found.')
            input(f'{WHITE} Press Enter to return to main menu...')
            sxr_main()
    filename = None
    if len(files) == 1:
        filename = files[0]
        print(f'{GREEN} [{RED}●{GREEN}] Found {len(files)} Excel Files:')
        for idx, f in enumerate(files, 1):
            print(f' {GREEN}[{RED}{idx}{GREEN}] {f}')
        print(LINE)
        choice = input(f'{GREEN} [{RED}●{GREEN}] Select File (1-{len(files)}) {EKL} ').strip()
        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(files):
                    filename = files[idx]
                    break
                    print(f'{RED} Invalid selection!')
                    continue
                    print(f'{GREEN} [{RED}●{GREEN}] Selected File {EKL} {filename}\n')
                    nums, err = extract_numbers_from_excel(filename)
                    if nums:
                        with open('Number_List.txt', 'w', encoding='utf-8', errors='ignore') as f:
                            for num in nums:
                                f.write(num + '\n')
                        print(f'{GREEN} [{RED}●{GREEN}] Success! Extracted {len(nums)} numbers From {filename} File.')
                        print(f'{GREEN} [{RED}●{GREEN}] Saved to \'Number_List.txt\'\n')
                        input(f'{WHITE} Press Enter to Start Forgetting {len(nums)} Numbers...')
                        autom_main()
                    else:
                        print(f'{RED} Error: {err}')
                        input(f'{WHITE} Press Enter to return Main Menu...')
                        sxr_main()
def autom_main():
    # irreducible cflow, using cdg fallback
    global CURRENT_LOCALE
    # ***<module>.autom_main: Failure: Compilation Error
    clear_logo()
    with open('Number_List.txt', 'r', encoding='utf-8', errors='ignore') as f:
        numbers = [line.strip() for line in f if line.strip()]
    if not numbers:
        print(f'{RED} No numbers found in Number_List.txt!')
        input(f'{WHITE} Press Enter to return to main menu...')
        sxr_main()
            return
        server_domain = SELECTED_SERVER
        print(f'{GREEN} [{RED}●{GREEN}] Server      {EKL} {server_domain}')
        print(f'{GREEN} [{RED}●{GREEN}] Device      {EKL} {SELECTED_DEVICE}')
        print(f'{GREEN} [{RED}●{GREEN}] Browser     {EKL} {SELECTED_BROWSER}')
        print(LINE)
        print(f'{WHITE} Setting up Main Proxy System...')
        PROXY_LIST = get_proxy_list('proxy_settings', 'Main Proxy')
        PROXY_ITERATOR = itertools.cycle(PROXY_LIST) if PROXY_LIST else None
        if PROXY_LIST:
            print(f'{GREEN} [{RED}●{GREEN}] Total Main Proxies {EKL} {len(PROXY_LIST)}')
        else:
            try:
                nfo = get_ip_info(None)
                loc = get_locale_code(nfo['countryCode'])
                CURRENT_LOCALE = loc
                print(f"{GREEN} [{RED}●{GREEN}] Direct Connection {EKL} {nfo['country']}")
                print(f'{GREEN} [{RED}●{GREEN}] Locale           {EKL} {loc}')
            except:
                CURRENT_LOCALE = 'en_US'
                print(f'{GREEN} [{RED}●{GREEN}] Direct Connection {EKL} No Proxy')
        print(LINE)
        print(f'{WHITE} Setting up SMS Proxy System...')
        SMS_PROXY_LIST = get_proxy_list('sms_proxy_settings', 'SMS Proxy')
        SMS_PROXY_ITERATOR = itertools.cycle(SMS_PROXY_LIST) if SMS_PROXY_LIST else None
        if SMS_PROXY_LIST:
            print(f'{GREEN} [{RED}●{GREEN}] Total SMS Proxies {EKL} {len(SMS_PROXY_LIST)}')
        else:
            print(f'{YELLOW} No SMS Proxy configured. Will use Main Proxy or Direct Connection')
        try:
            w_inp = input(f'{GREEN} [{RED}●{GREEN}] Enter number of Threads (Recommended: 10-30) {EKL} ').strip()
            if w_inp:
                max_workers = int(w_inp)
                if max_workers < 1:
                    max_workers = 10
                else:
                    if max_workers > 100:
                        max_workers = 50
            else:
                max_workers = 20
        except:
            max_workers = 20
        clear_logo()
        reset_counters()
        total_numbers = len(numbers)
        print(f'{GREEN} [{RED}●{GREEN}] Total Numbers {EKL} {total_numbers}')
        print(f'{GREEN} [{RED}●{GREEN}] Threads      {EKL} {max_workers}')
        print(f'{LINE}\n')
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = []
            for num in numbers:
                proxy_data = next(PROXY_ITERATOR) if PROXY_ITERATOR else None
                current_proxy = proxy_data['proxy'] if proxy_data else None
                current_locale = proxy_data['locale'] if proxy_data else CURRENT_LOCALE
                future = executor.submit(check, num, current_proxy, current_locale, SELECTED_BROWSER, 0, server_domain, SMS_PROXY_ITERATOR, SELECTED_DEVICE)
                futures.append(future)
            completed = 0
            for future in as_completed(futures):
                completed += 1
                try:
                    future.result(timeout=120)
                except Exception as e:
                    safe_print(f'{RED} Thread Error: {e}')
                with print_lock:
                    sys.stdout.write(f'\r{GREEN} Progress: {completed}/{total_numbers} numbers processed...')
                    sys.stdout.flush()
        with print_lock:
            sys.stdout.write('\r                                                                                \r')
            sys.stdout.flush()
        print(LINE)
        print(f'{GREEN} [{RED}●{GREEN}] {WHITE}Completed Forgetting {total_checked} Numbers.')
        print(f'{GREEN} [{RED}●{GREEN}] {GREEN}Total Success: {total_success} Numbers.')
        print(f'{GREEN} [{RED}●{GREEN}] {YELLOW}Total Failed: {total_failed} Numbers.')
        print(f'{GREEN} [{RED}●{GREEN}] {RED}Total Error: {total_error} Numbers.')
        print(LINE)
        try:
            choice = input(f'{WHITE} Press Enter to Start Again or Type \'M\' for Main Menu {EKL} ').strip().lower()
            if choice == 'm':
                sxr_main()
            else:
                if choice == '':
                    autom_main()
                else:
                    sxr_main()
        except Exception as e:
            print(f'{RED} Error: {e}')
            sxr_main()
            except FileNotFoundError:
                print(f'{RED} Number_List.txt file not found!')
                input(f'{WHITE} Press Enter to return to main menu...')
                sxr_main()
                    return None
                except Exception as e:
                        print(f'{RED} Error reading file: {e}')
                        input(f'{WHITE} Press Enter to return...')
                        sxr_main()
                            return None
def process_sms(session, resp_text, number, url, base_headers, server_domain, sms_proxy_iterator=None, device_type='Android'):
    # ***<module>.process_sms: Failure detected at line number 438 and instruction offset 18: Different bytecode
    if 'id=\"contact_point_selector_form\"' in resp_text and 'name=\"recover_method\"' in resp_text:
            sms_options = re.findall('input type=\"radio\" name=\"recover_method\" value=\"(send_sms:.*?)\".*?id=\"(.*?)\"', resp_text)
            target_value = None
            for val, inp_id in sms_options:
                label_match = re.search('label for=\"' + re.escape(inp_id) + '\".*?<div class=\"_52jc _52j9\">(.*?)</div>', resp_text, re.DOTALL)
                if label_match:
                    visible_text = label_match.group(1)
                    vis_digits = ''.join(filter(str.isdigit, visible_text))
                    if number.endswith(vis_digits):
                        target_value = val
                        safe_print(f'{CYAN} SMS Option Found {EKL} {visible_text}')
                        break
            if target_value:
                if sms_proxy_iterator:
                    try:
                        proxy_data = next(sms_proxy_iterator)
                        session.proxies.update(proxy_data['proxy'])
                        safe_print(f'{CYAN} Reloading Page...')
                        reload_headers = base_headers.copy()
                        reload_headers.update({'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8', 'referer': url})
                        reload_response = session.get(url, headers=reload_headers)
                        if reload_response.status_code == 200:
                            resp_text = reload_response.text
                        else:
                            safe_print(f'{RED} Page Reload Failed ({reload_response.status_code})')
                    except Exception as e:
                        safe_print(f'{RED} Proxy Switch/Reload Error: {e}')
                try:
                    lsd = re.search('name=\"lsd\" value=\"(.*?)\"', resp_text).group(1) if re.search('name=\"lsd\" value=\"(.*?)\"', resp_text) else ''
                    jazoest = re.search('name=\"jazoest\" value=\"(.*?)\"', resp_text).group(1) if re.search('name=\"jazoest\" value=\"(.*?)\"', resp_text) else ''
                    action_match = re.search('<form.*?action=\"(.*?)\".*?id=\"contact_point_selector_form\"', resp_text, re.DOTALL)
                    if action_match:
                        action_url = action_match.group(1).replace('&amp;', '&')
                        full_url = f'https://{server_domain}{action_url}'
                    else:
                        full_url = f'https://{server_domain}/ajax/recover/initiate/'
                    headers = base_headers.copy()
                    headers.update({'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8', 'cache-control': 'max-age=0', 'content-type': 'application/x-www-form-urlencoded', 'origin': f'https://{server_domain}', 'referer': url})
                    data = {'lsd': lsd, 'jazoest': jazoest, 'recover_method': target_value, 'reset_action': 'Continue'}
                    params = {'c': '/login/', 'ctx': 'initate_view', 'sr': '0', 'ars': 'facebook_login'}
                    sxr_respns = session.post(full_url, headers=headers, data=data, params=params)
                    if 'action=\"/recover/code/' in sxr_respns.text:
                        update_counter('success', number, 'SMS Sent Successfully', GREEN)
                        return True
                    else:
                        update_counter('failed', number, 'Code Sent Failed - Skipping...', RED)
                        return True
                except:
                    pass
            else:
                update_counter('failed', number, 'SMS Option Not Found/Mismatch - Skipping...', YELLOW)
                return True
    return False
def check(number, proxy=None, locale='en_US', browser_type='Chrome', retry_count=0, server_domain='limited.facebook.com', sms_proxy_iterator=None, device_type='Android'):
    # irreducible cflow, using cdg fallback
    # ***<module>.check: Failure: Compilation Error
    sxr_respns = None
    session = requests.Session()
    session.timeout = 30
    if proxy:
        session.proxies.update(proxy)
    else:
        if PROXIES:
            session.proxies.update(PROXIES)
    if device_type == 'Android':
        andro_ver = random.choice(['10', '11', '12', '13', '14'])
        models = ['SM-G998B', 'SM-S908B', 'Pixel 6', 'Pixel 7', 'Pixel 8', 'M2101K6G', '2201122G', '23127PN0CC']
        model = random.choice(models)
        chrome_ver = random.randint(90, 122)
        ua = f'Mozilla/5.0 (Linux; Android {andro_ver}; {model}) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{chrome_ver}.0.0.0 Mobile Safari/537.36'
    else:
        if device_type == 'iPhone':
            ios_ver = random.choice(['15_6_1', '16_0', '16_1', '16_2', '16_3', '16_4', '16_5', '17_0', '17_1'])
            iphone_model = random.choice(['iPhone13,2', 'iPhone14,3', 'iPhone14,5', 'iPhone15,2', 'iPhone15,4'])
            ios_main = ios_ver.split('_')[0]
            ua = f'Mozilla/5.0 (iPhone; CPU iPhone OS {ios_ver} like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/{ios_main}.0 Mobile/15E148 Safari/604.1'
        else:
            if device_type == 'KaiOS':
                kai_ver = random.choice(['2.5', '3.0', '3.1'])
                ua = f'Mozilla/5.0 (Mobile; Nokia; rv:48.0) Gecko/48.0 Firefox/48.0 KaiOS/{kai_ver}'
            else:
                if device_type == 'Windows Phone':
                    wp_ver = random.choice(['10.0', '8.1', '8.0'])
                    ua = f'Mozilla/5.0 (Windows Phone {wp_ver}; Android 7.0; Microsoft; Lumia 950) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Mobile Safari/537.36 Edge/15.15063'
                else:
                    if device_type == 'BlackBerry':
                        bb_ver = random.choice(['10.3.3', '10.3.2', '10.3.1'])
                        ua = f'Mozilla/5.0 (BB10; {bb_ver}) AppleWebKit/537.10+ (KHTML, like Gecko) Version/10.0.0.1337 Mobile Safari/537.10+'
                    else:
                        andro_ver = random.choice(['11', '12', '13'])
                        models = ['SM-G998B', 'Pixel 6', 'M2101K6G']
                        model = random.choice(models)
                        ua = f'Mozilla/5.0 (Linux; Android {andro_ver}; {model}) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/{random.randint(90, 120)}.0.0.0 Mobile Safari/537.36'
    base_headers = {'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8', 'accept-language': f'{locale},en;q=0.9', 'user-agent': ua, 'upgrade-insecure-requests': '1', 'sec-fetch-dest': 'document', 'sec-fetch-mode': 'navigate', 'sec-fetch-site': 'none', 'sec-fetch-user': '?1', 'cache-control': 'max-age=0'}
    if retry_count == 0:
        safe_print(f'{LIGHT_GRAY} Searching For {number}...')
    git_fb = session.get(f'https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0&__mmr=1&_rdr', headers=base_headers, timeout=30)
    try:
        lsd = re.search('name=\"lsd\" value=\"(.*?)\"', str(git_fb.text)).group(1)
    except:
        try:
            lsd = re.search('\\[\"LSD\",\\[\\],\\{\"token\":\"(.*?)\"\\}', str(git_fb.text)).group(1)
        except:
            lsd = ''
    try:
        jazoest = re.search('name=\"jazoest\" value=\"(.*?)\"', str(git_fb.text)).group(1)
    except:
        try:
            jazoest = re.search('\"initSprinkleValue\":\"(.*?)\"', str(git_fb.text)).group(1)
        except:
            jazoest = ''
    _data = {'lsd': lsd, 'jazoest': jazoest, 'email': number, 'did_submit': 'Search'}
    post_headers = base_headers.copy()
    post_headers.update({'content-type': 'application/x-www-form-urlencoded', 'origin': f'https://{server_domain}', 'referer': f'https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0', 'sec-fetch-site': 'same-origin'})
    url = f'https://{server_domain}/login/identify/?ctx=recover&c=%2Flogin%2F&search_attempts=1&ars=facebook_login&alternate_search=0'
    sxr_respns = session.post(url, data=_data, headers=post_headers, allow_redirects=True, timeout=30)
    if 'id=\"login_identify_search_error_msg\"' in sxr_respns.text:
        update_counter('failed', number, 'Account Not Found', MAGENTA)
        if 'action=\"/login/identify/?ctx=recover' in sxr_respns.text:
            update_counter('failed', number, 'Multiple Account Found - Skipping...', GOLD)
            if sxr_respns.url.startswith(f'https://{server_domain}/login/account_recovery/name_search/'):
                headers = base_headers.copy()
                headers.update({'referer': f'https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0&__mmr=1&_rdr', 'sec-fetch-site': 'same-origin'})
                sxr_respns = session.get(sxr_respns.url, headers=headers, timeout=30)
                safe_print(f'{VIOLET} Clicking Try to another way...')
                if 'action=\"/login/account_recovery/name_search/?flow=initiate_view' in sxr_respns.text:
                    headers = base_headers.copy()
                    headers.update({'referer': sxr_respns.url})
                    sxr_respns = session.get(f'https://{server_domain}/recover/initiate/?c=%2Flogin%2F&fl=initiate_view&ctx=msite_initiate_view', headers=headers, timeout=30)
                    if process_sms(session, sxr_respns.text, number, sxr_respns.url, base_headers, server_domain, sms_proxy_iterator, device_type):
                        return
                        if 'name=\"pass\"' in sxr_respns.text and '/login/account_recovery/' in sxr_respns.text:
                            update_counter('failed', number, 'Only Password Option Found - Skipping...', ORANGE)
                            return
                        else:
                            update_counter('error', number, 'Unknown Page (No Selector) - Skipping...', ORANGE, html_content=sxr_respns.text)
                if sxr_respns.url.startswith(f'https://{server_domain}/login/device-based/ar/login/?ldata='):
                    headers = base_headers.copy()
                    headers.update({'referer': f'https://{server_domain}/login/identify/?ctx=recover&ars=facebook_login&from_login_screen=0&__mmr=1&_rdr', 'sec-fetch-site': 'same-origin'})
                    sxr_respns = session.get(sxr_respns.url, headers=headers, timeout=30)
                    if 'id=\"contact_point_selector_form\"' in sxr_respns.text:
                        try:
                            try_another_way_url = re.search('href=\"(/recover/initiate/\\?privacy_mutation_token=.*?)\"', sxr_respns.text).group(1)
                            try_another_way_url = try_another_way_url.replace('&amp;', '&')
                        except:
                            try_another_way_url = '/recover/initiate/'
                        is_sms_checked = re.search('input type=\"radio\" name=\"recover_method\" value=\"send_sms:.*?\".*?checked=\"1\"', sxr_respns.text)
                        if is_sms_checked:
                            if process_sms(session, sxr_respns.text, number, sxr_respns.url, base_headers, server_domain, sms_proxy_iterator, device_type):
                                return
                                return None
                            headers = base_headers.copy()
                            headers.update({'referer': sxr_respns.url})
                            sxr_respns = session.get(f'https://{server_domain}{try_another_way_url}', headers=headers, timeout=30)
                            safe_print(f'{VIOLET} Clicking Try to another way...')
                            if process_sms(session, sxr_respns.text, number, sxr_respns.url, base_headers, server_domain, sms_proxy_iterator, device_type):
                                return
                                update_counter('error', number, 'Unknown Page after try another way - Skipping...', ORANGE, html_content=sxr_respns.text)
                        if 'name=\"captcha_response\"' in sxr_respns.text:
                            update_counter('failed', number, 'Captcha Found - Skipping...', PURPLE)
                            return
                        else:
                            if '/help/121104481304395' in sxr_respns.text or '/help/103873106370583' in sxr_respns.text:
                                update_counter('failed', number, 'Account Disabled - Skipping...', TOXIC)
                                return
                            else:
                                if 'class=\"area error\"' in sxr_respns.text:
                                    if retry_count < 3:
                                        check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator, device_type)
                                    return None
                                else:
                                    update_counter('error', number, 'Unknown Page (Device Based) - Skipping...', ORANGE, html_content=sxr_respns.text)
                    if 'window.MPageLoadClientMetrics' in sxr_respns.text:
                        if retry_count < 3:
                            check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator, device_type)
                        else:
                            update_counter('error', number, 'Unknown Page (Bot Block) - Skipping...', RED, html_content=sxr_respns.text)
                    else:
                        if '/r.php?next=' in sxr_respns.text or '/login.php?next=' in sxr_respns.text:
                            if retry_count < 3:
                                check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator, device_type)
                            return None
                        else:
                            if 'Your Request Couldn\'t be Processed' in sxr_respns.text:
                                update_counter('error', number, 'Your Request Couldn\'t be Processed', RED, html_content=sxr_respns.text)
                            else:
                                update_counter('error', number, 'Unknown Page - Skipping...', ORANGE, html_content=sxr_respns.text)
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout, requests.exceptions.ChunkedEncodingError) as e:
                    if retry_count < 3:
                        time.sleep(2)
                        check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator, device_type)
                        return None
                    else:
                        update_counter('error', number, 'Network Error - Skipping...', RED)
                except Exception as e:
                        if retry_count < 3:
                            check(number, proxy, locale, browser_type, retry_count + 1, server_domain, sms_proxy_iterator, device_type)
                            return None
                        else:
                            update_counter('error', number, 'Error - Skipping...', RED)
def sxr_secure_start():
    """Approval Check করে তারপর Tool চালায়"""
    global settings
    settings = load_settings()
    if check_approval():
        apvv()
if __name__ == '__main__':
    sxr_secure_start()
